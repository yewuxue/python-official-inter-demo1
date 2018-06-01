#!/usr/bin/env python
# -*- coding: utf-8 -*-
#Author:YangShuang

from openpyxl import load_workbook
from config.read_excel import read_Excel

class write_excel(object):
    '''修改excel数据'''
    def __init__(self, filePath,Sheet):
        self.filename = filePath
        self.wb = load_workbook(self.filename)
        #self.ws = self.wb.active  # 获取当前活跃的Worksheet
        #self.ws = self.wb["Sheet1"]
        self.ws = self.wb[Sheet]

    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)

if __name__ == "__main__":
    filePath = r"D:\pycharmworkspace\demo1\testdata.xlsx"
    sheetName = "Sheet2"
    we=write_excel(filePath,sheetName)
    we.write(4,5,"qwer")

